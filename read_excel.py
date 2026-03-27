import openpyxl
import json

wb = openpyxl.load_workbook('bibliotek-sn.xlsx')
output = {}
output['sheets'] = wb.sheetnames

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    output[sheet_name] = {
        'rows': ws.max_row,
        'cols': ws.max_column,
        'data': []
    }
    for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True):
        output[sheet_name]['data'].append([str(c) if c is not None else None for c in row])

with open('excel_output.json', 'w') as f:
    json.dump(output, f, indent=2, ensure_ascii=False)

print("Done - wrote excel_output.json")
